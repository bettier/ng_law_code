import xlrd 
import xlwt 
import xlsxwriter
import os 
import re #regularization
from openpyxl import load_workbook 

def process(read, write): 
	readRow = read.nrows 

	for i in range(1, readRow): 
		plantiff_reason = read.cell(i, 22).value 
		reasons_for = read.cell(i, 27).value 
		both = read.cell(i, 14).value

		s = 'Z'+str(i+1)
		if re.search(r'分居', plantiff_reason) or re.search(r'分居', reasons_for): 
			write[s] = "y"
		elif re.search(r'n',both): 
			write[s] = "n/a"
		else: 
			write[s] = "n"

		d = 'Y'+str(i+1)
		if re.search(r'被告[\u4e00-\u9fa5]*同意|被告[\u4e00-\u9fa5]*自愿',reasons_for): 
			write[d] = "y"
		elif re.search(r'n',both): 
			write[d] = "n/a"
		else: 
			write[d] = "n"

if __name__ == '__main__':
	path = "../../Downloads/"
	final = "2.1 贵阳市all case.xlsx"

	dest = path+final 

	toRead = xlrd.open_workbook(dest)
	read = toRead.sheets()[0]

	target1 = load_workbook(dest)
	ws = target1['贵阳市']
	process(read, ws)
	target1.save(dest)

