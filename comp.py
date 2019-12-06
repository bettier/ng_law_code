import xlrd 
import xlwt 
import xlsxwriter
import os 
import re #regularization
from openpyxl import load_workbook 

#def jufaComp(read, revise, jufa):
	#NUM3 

def wenshu(read, write, wenshu): 
	wenshuRow = wenshu.nrows
	readRow = read.nrows 

	read_caseNo = read.col_values(3, start_rowx=1, end_rowx=readRow)
	wenshu_caseNo = wenshu.col_values(0, start_rowx=0,end_rowx=wenshuRow)

	read_l = []
	wenshu_l = []

	for r in read_caseNo: 
		f = re.search(r'[\d]*号',r)
		num = ""
		if f: 
			num = r[f.start():f.end()]
		else: 
			if r != "": 
				print(r)
		read_l.append(num)

	for w in wenshu_caseNo: 
		f = re.search(r'[\d]*号',w)
		num = ""
		if f: 
			num = w[f.start():f.end()]
		else: 
			print(w)
		wenshu_l.append(num)

	not_in_read = []
	l_not = []
	for i in range(0,wenshuRow): 
		w = wenshu_l[i]
		if read_l.count(w) > 0: 
			pos = [i for i, n in enumerate(read_l) if n == w]
			for p in pos:
				new = p + 2 
				c = 'F'+str(new) 
				write[c] = "wenshu"
		else: 
			not_in_read.append(w)
			l_not.append(i)
	return not_in_read, l_not 

def printInvalid(sheet):
	ncols = sheet.ncols
	nrows = sheet.nrows
	print("the rows: ", nrows)
	for i in range(0,nrows): 
		to = i + 2 
		if sheet.cell(i,0).value == "": 
			print(to)

if __name__ == '__main__':
	path = "../../Downloads/"
	final = "4.2 哈尔滨市case2015年s.xlsx"

	dest = path + final 

	print("THE START OF READING")
	
	toRead = xlrd.open_workbook(dest)
	read = toRead.sheets()[0]

	wenshu1 = xlrd.open_workbook("./2015H.xls")
	wenshu_s = wenshu1.sheets()[0]

	target1 = load_workbook(dest)
	ws = target1['#01-50']
	#ws['A1'] = "TRY"
	
	l_index, n = wenshu(read, ws, wenshu_s)
	target1.save(dest)
	#xlrd.open_workbook(dest)
	print("FINISH READING")
	print(l_index)
	print("THERE ARE: ", len(l_index))
	print(n)
	#printInvalid(read)

	#wb = copy(target1)
	#open_sheet = target1.sheet_by_index(0)
	#wb = copy(open_sheet)
	#open_sheet.write(0,0,"TRY")
	#wb.save(dest)
	
