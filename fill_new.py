import xlrd 
import xlwt 
import xlsxwriter
import os 
import re #regularization
from openpyxl import load_workbook 

#def addJufa(): 

def addWenshu(wenshu,jufa, jufa2, read, write, j_s, w_s):
	wenshuRow = wenshu.nrows
	jufaRow = jufa.nrows
	
	jufaCol = jufa.ncols
	#print("The NUMBER IS:", jufaCol)
	jufa2Row = jufa2.nrows
	readRow = read.nrows 

	wenshu_caseNo = wenshu.col_values(0, start_rowx=0,end_rowx=wenshuRow)
	jufa_caseNo = jufa.col_values(10, start_rowx=1,end_rowx=jufaRow)
	jufa2_caseNo = jufa2.col_values(10, start_rowx=1,end_rowx=jufa2Row)
	read_caseNo = read.col_values(3, start_rowx=1, end_rowx=readRow)
	
	wenshu_l = []
	jufa_l = []
	jufa2_l = []
	read_l = []

	j_count = 0 
	w_count = 0 
	for r in wenshu_caseNo: 
		f = re.search(r'[\d]*号',r)
		num = ""
		if f: 
			num = r[f.start():f.end()]
		else: 
			if r != "": 
				print(r)
		wenshu_l.append(num)

	for r in jufa_caseNo: 
		f = re.search(r'[\d]*号',r)
		num = ""
		if f: 
			num = r[f.start():f.end()]
		else: 
			if r != "": 
				print(r)
		jufa_l.append(num)

	for r in jufa2_caseNo: 
		f = re.search(r'[\d]*号',r)
		num = ""
		if f: 
			num = r[f.start():f.end()]
		else: 
			if r != "": 
				print(r)
		jufa2_l.append(num)

	for r in read_caseNo: 
		f = re.search(r'[\d]*号',r)
		num = ""
		if f: 
			num = r[f.start():f.end()]
		else: 
			if r != "": 
				print(r)
		read_l.append(num)

	#add the new stuff into the sheet 
	index = 1
	for j in jufa_l: 
		rC = read_l.count(j) 
		wC = wenshu_l.count(j)

		if wC == 0 and rC > 0: 
			pos = [i for i, n in enumerate(read_l) if n == j]
			for p in pos:
				new = p + 2 
				c = 'F'+str(new)
				if read.cell(new, 5).value != "wenshu":  
					write[c] = "jufa"
				else:
					print("The final: ", read.cell(new, 3).value, "the jufa: ", j)
		elif wC > 0 and rC == 0:
			the_row = jufa.row_values(index, start_colx=0, end_colx=jufaCol)
			for st in range(0, jufaCol): 
				w_s.write(w_count, st, the_row[st])
			w_count = w_count + 1 
		elif wC == 0 and rC == 0: 
			the_row = jufa.row_values(index, start_colx=0, end_colx=jufaCol)
			for st in range(0, jufaCol): 
				j_s.write(j_count, st, the_row[st])
			j_count = j_count + 1

		index = index + 1

	index2 = 1
	for j in jufa2_l: 
		rC = read_l.count(j) 
		wC = wenshu_l.count(j)

		if wC == 0 and rC > 0: 
			pos = [i for i, n in enumerate(read_l) if n == j]
			for p in pos:
				new = p + 2 
				c = 'F'+str(new)
				if read.cell(new, 5).value != "wenshu":  
					write[c] = "jufa"
				else:
					print("The final: ", read.cell(new, 3).value, "the jufa: ", j)
					
		elif wC > 0 and rC == 0:
			the_row = jufa2.row_values(index2, start_colx=0, end_colx=jufaCol)
			for st in range(0, jufaCol): 
				w_s.write(w_count, st, the_row[st])
			w_count = w_count + 1 
		elif wC == 0 and rC == 0: 
			the_row = jufa2.row_values(index2, start_colx=0, end_colx=jufaCol)
			for st in range(0, jufaCol): 
				j_s.write(j_count, st, the_row[st])
			j_count = j_count + 1

		index2 = index2 + 1 

	for w in wenshu_l: 
		if read_l.count(w) == 0 and jufa_l.count(w) == 0 and jufa2_l.count(w) == 0: 
			w_s.write(w_count,10, w) 
			w_count = w_count + 1 

if __name__ == '__main__':
	path = "../../Downloads/"
	final = "2.1 贵阳市all case.xlsx"
	jufa14 = "jufa贵阳2014.xlsx"
	jufa15 = "jufa贵阳2015.xlsx"

	dest = path + final 
	jufaD14 = path + jufa14
	jufaD15 = path + jufa15 

	print("THE START OF READING")
	
	toRead = xlrd.open_workbook(dest)
	read = toRead.sheets()[0]

	wenshu1 = xlrd.open_workbook("./2014GY.xls")
	wenshu_s = wenshu1.sheets()[0]

	jufa14 = xlrd.open_workbook(jufaD14)
	jufa_s14 = jufa14.sheets()[0]
	jufa15 = xlrd.open_workbook(jufaD15)
	jufa_s15 = jufa15.sheets()[0]

	target1 = load_workbook(dest)
	ws = target1['贵阳市']

	write_to = xlwt.Workbook() 
	jufa_sheet = write_to.add_sheet(u'jufa', cell_overwrite_ok=True)
	wenshu_sheet = write_to.add_sheet(u'wenshu', cell_overwrite_ok=True)
	addWenshu(wenshu_s, jufa_s14, jufa_s15, read, ws, jufa_sheet, wenshu_sheet)

	target1.save(dest)
	write_to.save("./toProcessedGY.xls")
